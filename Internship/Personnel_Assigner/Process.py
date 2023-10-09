import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import Calendar
from datetime import date
import pandas as pd
import openpyxl
from babel import  numbers # needed for executable generation

""""

This script has been developed for the manual excel file "Test.xlsx" which has different parameter names then the official resource file. 

Check Process_test_v2 for implementation with official given files.

Steps include:
- Reading the excel file only of the personnel information
- Storing the data from excel to a list of dictionaries (each dictionary corresponding to one set of personnel information for each personnel)
    - Later implementations might proof that a database implementation should be used etc.
- UI is formed including the implementation of the "Process & Filter" of the information gathered both from the read excel file and user input via the UI

Use collect_info from MainApplication for process implementation

For further documentation and explanation please check the related parts
"""

def count_non_empty_rows(file_path, sheet_name):
    """

    Counts the full rows of a excel file so that while reading that file it is possible to adjust the range.

    :param file_path:   Excel File to be read
    :param sheet_name:  Sheet name in the Excel File that has the lines to be read
    :return:            Number of full lines with entry
    """
    try:
        workbook = openpyxl.load_workbook(file_path)    # Opens the Excel File into an object
        sheet = workbook[sheet_name]                    # Assigns the particular sheet to an object for easier operation
        row_count = 0                                   # Initializes the counter variable with 0

        for row in sheet.iter_rows():                   # iter_rows() returns an iterator (which is a Python object) in this context, for each iteration, a tuple (containing cells of a particular row) is returned
            if any(cell.value for cell in row):         # row is a tuple of cells in a row per iteration; thus, for every cell in it is check whether if it stores any value in it, if so, any() return True, if not False
                row_count += 1                          # Since any() return True for cells with value in it, that row must be full; thus, counter goes up

        return row_count                                # No exception -> return counter value
    except Exception as e:                              # Exception is type and contains a wide range of Exceptions, e is the particular instance
        print(f"Error: {e}")                            # since e is the instance, it indicates the error/exception that as occured
        return None

def convert_list(b_converted):

    """
    Takes list a list of dictionaries and merges the same ones and counts them at the same time, adds a "count" key to each dictionary.
    This is necessary to use the "filter_personnel()" function due to it's parameter format for required_personnel. Input to this function is generally
    the user input coming through the UI

    Inputed list format:
    input = [
        {"Org Unit": "U1", "Pos": "M Op"},
        {"Org Unit": "U1", "Pos": "M Op"},
        {"Org Unit": "U1", "Pos": "F Op"},
        {"Org Unit": "U2", "Pos": "M Op"}
    ]
    Returned list format:
    output = [
        {"Org Unit": "U1", "Pos": "M Op", "Count": 2},
        {"Org Unit": "U1", "Pos": "F Op", "Count": 1},
        {"Org Unit": "U2", "Pos": "M Op", "Count": 1}
    ]

    :param b_converted: A list of dictionaries with "Organization Unit" and "Position" (other than Personnel Number, Names and Surnames)
    :return: A list of dictionaries with it's dictionaries with the same "Organization Unit" and "Position" merged and counted, adding a "Count" parameter aswell
    """

    # Add a count of 1 for each element initially
    for item in b_converted:
        item["Count"] = 1

    # Create a dictionary to keep track of merged items, ie. hold them temporarily
    merged_dict = {}

    # Holds one key in key variable to be compared with merged_dict which always holds the cumulative of the prior dictionaries
    # so if the key is there, counter goes up for that dictionary, if not key also is added to the merged_dict (which is also a dictionary of dictionaries)
    for item in b_converted:
        key = (item["Organization Unit"], item["Position"])
        if key in merged_dict:
            merged_dict[key]["Count"] += 1
        else:
            merged_dict[key] = item

    # Convert the merged dictionary back to a list since merged_dict was a dictionary
    merged_list = list(merged_dict.values())

    return  merged_list

def filter_personnel(personnel_list, required_personnel): #personnel_list = employees & required_personnel is our input
    """

    example for personnel_list:
    personnel_list:[
        {'Personal Num': 1, 'Name': 'A', 'Surname': 'F', 'Org Unit': 'U1', 'Pos': 'M Op'}
        {'Personal Num': 2, 'Name': 'B', 'Surname': 'G', 'Org Unit': 'U2', 'Pos': 'M Op'}
        {'Personal Num': 3, 'Name': 'C', 'Surname': 'H', 'Org Unit': 'Mech', 'Pos': 'F Op'}
        {'Personal Num': 4, 'Name': 'D', 'Surname': 'J', 'Org Unit': 'U1', 'Pos': 'F Op'}
        {'Personal Num': 5, 'Name': 'E', 'Surname': 'K', 'Org Unit': 'U2', 'Pos': 'H Op'}
        {'Personal Num': 6, 'Name': 'L', 'Surname': 'P', 'Org Unit': 'U1', 'Pos': 'M Op'}
    ]

    :param personnel_list: The list that has the information of all personnel (with names)
    :param required_personnel: The list that contains the information of wanted personnel (without names)
    :return: list of the wanted personnel with their names found from personnel_list

    """

    filtered_personnel = []
    # Criteria is renewed for each iteration
    for criteria in required_personnel:
        org_unit = criteria["Organization Unit"]
        pos = criteria["Position"]
        count = criteria["Count"]

        # New list that includes the personnel with fitting keys, this is a "comprehension list" syntax
        matching_personnel = [person for person in personnel_list if
                              person["Organization Unit"] == org_unit and person["Position"] == pos]

        # Since we give the count as an input, it is expected that to have that amount of matching personnel
        if len(matching_personnel) >= count:
            filtered_personnel.extend(matching_personnel[:count]) # The matching personel so far added to the filtered personnel list
        else:
            # If not enough personnel are found, we include all available personnel for that criteria
            filtered_personnel.extend(matching_personnel)
            print(
                f"Warning: Not enough personnel found for Org Unit '{org_unit}' and Position '{pos}'. Required: {count}, Found: {len(matching_personnel)}")

    return filtered_personnel




#############################################################   READ DATA FROM EXCEL    ################################################################

data_frame = pd.read_excel("Test.xlsx")

row_count = count_non_empty_rows("Test.xlsx", "Sheet1")

# Creating the dictionary array and filling it with information
employees = []
for i in range(0,row_count-1):
    employees.append({"Personal Num":data_frame.iat[i,0],
                      "Name":data_frame.iat[i,1],
                      "Surname":data_frame.iat[i,2],
                      "Organization Unit":data_frame.iat[i,3],
                      "Position":data_frame.iat[i,4]})
    #print(employees[i])

##############################################################  READ DATA FROM EXCEL    #########################################################



##############################################################      UI      ######################################################################
class MainApplication(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Assigner_demo")

        window_width = 700
        window_height = 500

        # Get the screen width and height
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Calculate the center coordinates
        x_coordinate = (screen_width - window_width) // 2
        y_coordinate = (screen_height - window_height) // 2

        self.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

        self.label1 = tk.Label(self, text="Enter the amount of personnel:")
        self.label1.pack(pady=10)

        self.entry_objects = tk.Entry(self)
        self.entry_objects.pack(pady=5)

        self.label2 = tk.Label(self, text="Select the date:")
        self.label2.pack(pady=10)

        self.calendar_var = tk.StringVar(self, value=date.today().strftime('%d.%m.%Y'))
        self.calendar = Calendar(self, selectmode='day', date_pattern='dd.mm.yyyy', textvariable=self.calendar_var)
        self.calendar.pack(pady=5)

        self.label3 = tk.Label(self, text="Select the time:")
        self.label3.pack(pady=10)

        self.time_var = tk.StringVar(self)
        self.time_var.set("00:00")

        self.time_dropdown = ttk.Combobox(self,
                                          values=[f"{h:02d}:{m:02d}" for h in range(24) for m in range(0, 60, 15)],
                                          textvariable=self.time_var)
        self.time_dropdown.pack(pady=5)

        self.collect_button = tk.Button(self, text="Proceed", command=self.collect_info)
        self.collect_button.pack(pady=20)

        self.attributes("-topmost", False)

        self.object_info_list = []

    def collect_info(self):
        num_objects = int(self.entry_objects.get())
        date_hour = f"{self.calendar_var.get()} {self.time_var.get()}"

        if num_objects <= 0:
            messagebox.showinfo("Warning", "Selected amount is 0.")
            return

        for object_number in range(1, num_objects + 1):
            popup = PopupWindow(self, date_hour, object_number)
            self.wait_window(popup)

        #################################################################         PROCESS & FILTER & OUTPUT         ########################################################
        # Variables for use:
        # self.object_info_list (Includes Organization Unit and Position)
        # date_hour
        # object_number

        # Count the dictionaries (personnel) with the same criteria in self.object_info_list

        #for input in self.object_info_list:

        # Create another criteria list formed of dictionaries (criterias)

        object_ready_list = convert_list(self.object_info_list) # Convert input into counted and merged list
        filtered_list = filter_personnel(employees, object_ready_list) # Filter the counted and merged list by comparing it to the employee list from Excel

        eligible_names = [(person["Name"], person["Surname"]) for person in filtered_list] # find Names

        print(eligible_names)

        print("\nEligible Personnel:")
        for name, surname in eligible_names:
            print(f"{name} {surname}")

        ##################################       OUTPUT      ########################################

        window_width = 300
        window_height = 300

        global x_offset, y_offset

        # Get the screen width and height
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Calculate the center coordinates
        x_coordinate = (screen_width - window_width) // 2
        y_coordinate = (screen_height - window_height) // 2

        x_shift = 50 + x_offset
        y_shift = 50 + y_offset

        window = tk.Tk()
        window.title("List Display")

        listbox = tk.Listbox(window, width=30, height=10)

        window.geometry(f"{window_width}x{window_height}+{x_coordinate + x_shift}+{y_coordinate + y_shift}")

        for item in eligible_names:
            # Combine the two words in each tuple and add them to the listbox
            combined_item = f"{item[0]} {item[1]}"
            listbox.insert(tk.END, combined_item)

        scrollbar = tk.Scrollbar(window)
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)

        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        window.mainloop()

        ##################################       OUTPUT      ########################################



"""
        personnel_counter = 0
        for personnel in employees:
            personnel_counter += 1
            if(self.object_info_list[personnel_counter]["Organization Unit" == employees[]]):
        print(self.object_info_list[0]["Organization Unit"])
"""
        #################################################################         PROCESS & FILTER & OUTPUT        ########################################################
        #print(date_hour)
        #print(object_number)
        #for obj_info in self.object_info_list:
            #print(obj_info)


class PopupWindow(tk.Toplevel):
    def __init__(self, master, date_hour, object_number):
        super().__init__(master)
        self.title(f"Personnel {object_number} Information Collection")

        window_width = 300
        window_height = 300

        global x_offset, y_offset

        # Get the screen width and height
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Calculate the center coordinates
        x_coordinate = (screen_width - window_width) // 2
        y_coordinate = (screen_height - window_height) // 2

        x_shift = 50 + x_offset
        y_shift = 50 + y_offset
        self.geometry(f"{window_width}x{window_height}+{x_coordinate + x_shift}+{y_coordinate + y_shift}")
        self.date_hour = date_hour
        self.object_number = object_number

        ############# First dropdown menu
        self.label1 = tk.Label(self, text="Choose Organization Unit:")
        self.label1.pack(pady=10)

        self.option_var1 = tk.StringVar(self)
        self.option_var1.set("U1")

        self.option_menu1 = tk.OptionMenu(self, self.option_var1, "U1", "U2", "Mech")
        self.option_menu1.pack(pady=5)
        ######################

        ########## Second dropdown menu
        self.label2 = tk.Label(self, text="Choose Position:")
        self.label2.pack(pady=10)

        self.option_var2 = tk.StringVar(self)
        self.option_var2.set("M Op")

        self.option_menu2 = tk.OptionMenu(self, self.option_var2, "M Op", "F Op", "H Op")
        self.option_menu2.pack(pady=5)
        #########################


        self.collect_button = tk.Button(self, text="Confirm", command=self.collect_and_close)
        self.collect_button.pack(pady=20)

        self.lift()
        x_offset += 10
        y_offset += 10

    def collect_and_close(self):
        chosen_option1 = self.option_var1.get()
        chosen_option2 = self.option_var2.get()


        obj_info = {
            "Organization Unit": chosen_option1,
            "Position": chosen_option2
        }

        self.master.object_info_list.append(obj_info)

        self.withdraw()
        messagebox.showinfo("Saved",
                            f"Date & Hour: {self.date_hour}\nOrganization Unit: {chosen_option1}\nPosition: {chosen_option2}")
        self.deiconify()
        self.destroy()


if __name__ == "__main__":
    x_offset = 0
    y_offset = 0
    app = MainApplication()
    app.mainloop()
###################################################################         UI          ####################################################